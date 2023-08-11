import os
import os.path
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
print("#-------------------------------------------#")
print("|               =====  =====                |")
print("|                 =    =====                |")
print("|               ===    =====                |")
print("|                                           |")
print("|           ╔══ ╖ ╓ ╔══ ╔══ ═╦═             |")
print("|           ╚═╗ ╠═╣ ╠══ ╠══  ║              |")
print("|           ══╝ ╜ ╙ ╚══ ╚══  ╨              |")
print("|         ╔╦╗ ╔══ ╔══ ╔═╗ ╔══ ╔══           |")
print("|         ║║║ ╠══ ║   ╚═╣ ╠══ ║             |")
print("|         ╜╨╙ ╚══ ╜   ══╝ ╚══ ╜             |")
print("|                  V0.1                     |")
print("#-------------------------------------------#")
global tableColumnLength
tableColumnLength = []
MergedWorkbook = openpyxl.Workbook()
MergedSheet = MergedWorkbook.active
print("Usasge: ")
print("1. Place this script inside a folder with\n   the worksheets.")
print("2. Input a master worksheet filename, this worksheet\n   will become the template for the other worksheets,\n   if all of your worksheets have different headers\n   then create one with all of the headers required\n   data inside headers not included will be discarded")
def getWorksheetTemplate(tableColumnLength):
    worksheetTemplateName = str(input())
    if len(worksheetTemplateName.split(".")) == 1: #file extension was not provided
        worksheetTemplateName = worksheetTemplateName + ".xlsx"
    if os.path.isfile(worksheetTemplateName):
        print("file found!")
        worksheetTemplate = openpyxl.load_workbook(worksheetTemplateName)
        sheetTemplate = worksheetTemplate.active # get first worksheet
        for char in sheetTemplate.tables.items()[0][1]: #find amount of columns in the template table by filtering for the letters
            if char.isalpha(): ##check if character is a number or not
                tableColumnLength.append(ord(char.lower())-96)
        #style = TableStyleInfo(name=sheetTemplate.tables.items()[0][0], ref=tableColumnLength)
        
        print(tableColumnLength)
        #for column in range(1, sheet.max_column + 1):
            
        
    else:
        print("Error: File not found, please enter correct filename")
        getWorksheetTemplate()

getWorksheetTemplate(tableColumnLength)
##tableRowPosition = 0
##        tempPositionNumber = ""
##        for char in sheetTemplate.tables.items()[0][1].split(":"):
##            if char.isdigit():
##                tempPositionNumber.append(char)
##        tableRowPosition = int(tempPositionNumber)
##        for char in sheetTemplate.tables.items()[0][1]: #find amount of columns in the template table by filtering for the letters
##            if char.isalpha(): ##check if character is a number or not
##                tableColumnLength.append(ord(char.lower())-96)
##        print("colum length:",tableColumnLength)
##        global style
##        #headerPosition = sheetTemplate.tables.items()[0][1].split(":")[0] + tableColumnLength
##        tablePositions = chr(tableColumnLength[0]).upper() + chr(tableRowPosition)+":"+chr(tableColumnLength[1]).upper() + chr(tableRowPosition)
##        tab = Table(displayName=sheetTemplate.tables.items()[0][0], ref=tablePositions)
